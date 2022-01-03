from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('sa-voluteer.xlsx')
ws = wb.active

for row in range(1, 11): # Main row
    for col in range(1, 5): # nested to access columns 1-4
        char = get_column_letter(col)
        print(ws[char + str(row)].value, ", ")

#todo think of a better way to format
       #  This way
       # char = chr(65 + col) # cell refrence for int but you need A1 A2 B3 and so on
       #  # gives the char that is representive of a int 65+col0 = A1, 65+col1 = B1
       #



